import io
from xml.etree import ElementTree as ET

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


def decode_text(content: bytes) -> str:
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return content.decode("gbk")
        except UnicodeDecodeError:
            return content.decode("utf-8", errors="ignore")


def extract_pdf_text(raw: bytes) -> str:
    reader = PdfReader(io.BytesIO(raw))
    parts: list[str] = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n\n".join(parts).strip()


def extract_docx_text(raw: bytes) -> str:
    doc = DocxDocument(io.BytesIO(raw))
    parts = [para.text for para in doc.paragraphs if para.text.strip()]
    return "\n".join(parts).strip()


def extract_xlsx_text(raw: bytes) -> str:
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if cell is None else str(cell) for cell in row]
            if any(cell.strip() for cell in cells):
                lines.append("\t".join(cells))
    workbook.close()
    return "\n".join(lines).strip()


def extract_xml_text(raw: bytes) -> str:
    root = ET.fromstring(raw)
    return ET.tostring(root, encoding="unicode")


def extract_text_content(raw: bytes, ext: str) -> str:
    normalized = ext.lower()
    if normalized in {".txt", ".md", ".csv", ".json", ".log"}:
        return decode_text(raw).strip()
    if normalized == ".xml":
        try:
            return extract_xml_text(raw)
        except ET.ParseError:
            return decode_text(raw).strip()
    if normalized == ".pdf":
        return extract_pdf_text(raw)
    if normalized == ".docx":
        return extract_docx_text(raw)
    if normalized == ".xlsx":
        return extract_xlsx_text(raw)

    text = decode_text(raw).strip()
    if text:
        return text
    raise ValueError(f"Unable to extract text from {normalized or 'unknown'} file")
